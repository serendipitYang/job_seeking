"""
Excel output generator for job search results.
"""

import os
from datetime import datetime
from typing import List, Dict, Optional
import logging

logger = logging.getLogger(__name__)


def generate_excel_output(
    jobs: List,
    similarity_results: List,
    config: Dict,
    output_dir: str = "output"
) -> str:
    """
    Generate Excel output with job listings and resume recommendations.

    Args:
        jobs: List of JobPosting objects
        similarity_results: List of SimilarityResult objects (parallel to jobs)
        config: Configuration dictionary
        output_dir: Output directory

    Returns:
        Path to generated Excel file
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        logger.error("Required packages not installed. Run: pip install pandas openpyxl")
        raise

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename_prefix = config.get("output", {}).get("filename_prefix", "job_matches")
    output_path = os.path.join(output_dir, f"{filename_prefix}_{timestamp}.xlsx")

    # Prepare data
    data = []
    for i, job in enumerate(jobs):
        similarity = similarity_results[i] if i < len(similarity_results) else None

        row = {
            "Company": job.company,
            "Job Title": job.title,
            "Location": job.location,
            "Posted Date": job.posted_date.strftime("%Y-%m-%d") if job.posted_date else "Unknown",
            "Job URL": job.url,
            "Job ID": job.job_id,
        }

        if similarity:
            row["Recommended Resume"] = similarity.recommended_resume
            row["Match Score"] = f"{similarity.recommended_score:.1%}"
            row["All Scores"] = similarity.all_scores_display
        else:
            row["Recommended Resume"] = "N/A"
            row["Match Score"] = "N/A"
            row["All Scores"] = "N/A"

        data.append(row)

    # Create DataFrame
    df = pd.DataFrame(data)

    # Sort by match score (descending) then by posted date (most recent first)
    if "Match Score" in df.columns and len(df) > 0:
        df["_sort_score"] = df["Match Score"].apply(
            lambda x: float(x.rstrip("%")) / 100 if x != "N/A" else 0
        )
        df = df.sort_values(["_sort_score", "Posted Date"], ascending=[False, False])
        df = df.drop("_sort_score", axis=1)

    # Create Excel workbook with styling
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Matches"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # High match highlight
    high_match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    medium_match_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Write headers
    headers = list(df.columns) if len(df) > 0 else [
        "Company", "Job Title", "Location", "Posted Date", "Job URL",
        "Job ID", "Recommended Resume", "Match Score", "All Scores"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row_data in enumerate(df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Highlight based on match score
            if headers[col_idx - 1] == "Match Score" and value != "N/A":
                try:
                    score = float(value.rstrip("%")) / 100
                    if score >= 0.5:
                        ws.cell(row=row_idx, column=col_idx).fill = high_match_fill
                    elif score >= 0.35:
                        ws.cell(row=row_idx, column=col_idx).fill = medium_match_fill
                except:
                    pass

            # Make URLs clickable
            if headers[col_idx - 1] == "Job URL" and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

    # Adjust column widths
    column_widths = {
        "Company": 20,
        "Job Title": 40,
        "Location": 25,
        "Posted Date": 12,
        "Job URL": 50,
        "Job ID": 15,
        "Recommended Resume": 20,
        "Match Score": 12,
        "All Scores": 50,
    }

    for col, header in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = column_widths.get(header, 15)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")

    summary_data = [
        ["Job Search Summary", ""],
        ["", ""],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Days Lookback", config.get("days_lookback", 7)],
        ["Total Jobs Found", len(jobs)],
        ["", ""],
        ["Jobs by Company", ""],
    ]

    # Count by company
    if len(df) > 0:
        company_counts = df["Company"].value_counts().to_dict()
        for company, count in company_counts.items():
            summary_data.append([company, count])

    summary_data.extend([
        ["", ""],
        ["Jobs by Recommended Resume", ""],
    ])

    # Count by recommended resume
    if len(df) > 0 and "Recommended Resume" in df.columns:
        resume_counts = df["Recommended Resume"].value_counts().to_dict()
        for resume, count in resume_counts.items():
            summary_data.append([resume, count])

    for row_idx, row in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif "by Company" in str(row[0]) or "by Recommended" in str(row[0]):
                cell.font = Font(bold=True)

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 15

    # Save workbook
    wb.save(output_path)
    logger.info(f"Excel output saved to: {output_path}")

    return output_path


def load_applied_companies(excel_path: str) -> List[str]:
    """
    Load company names from the applied positions Excel file.

    Args:
        excel_path: Path to the Excel file with applied positions

    Returns:
        List of unique company names
    """
    try:
        import pandas as pd
    except ImportError:
        logger.error("pandas not installed")
        return []

    if not os.path.exists(excel_path):
        logger.warning(f"Applied positions file not found: {excel_path}")
        return []

    companies = set()
    try:
        xlsx = pd.ExcelFile(excel_path)
        for sheet in xlsx.sheet_names:
            df = pd.read_excel(xlsx, sheet_name=sheet)
            if "公司" in df.columns:
                sheet_companies = df["公司"].dropna().unique().tolist()
                companies.update(sheet_companies)
                logger.info(f"Loaded {len(sheet_companies)} companies from sheet '{sheet}'")
    except Exception as e:
        logger.error(f"Error reading applied positions file: {e}")

    return list(companies)
