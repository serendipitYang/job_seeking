#!/usr/bin/env python3
"""
Job Seeking Tool - Main Entry Point

Searches for intern/co-op positions at major tech companies,
matches them against your resumes, and generates an Excel report.

Usage:
    python run_job_search.py [--days N] [--config PATH] [--no-similarity]
"""

import os
import sys
import argparse
import logging
from datetime import datetime
from pathlib import Path

import yaml

# Add src to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "src"))

from scrapers import fetch_all_jobs, JobPosting
from similarity import create_matcher, SimilarityResult
from output import generate_excel_output, load_applied_companies
from company_discovery import load_companies_from_excel, generate_company_configs
import pandas as pd

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("job_search.log"),
    ]
)
logger = logging.getLogger(__name__)


def load_config(config_path: str = "config.yaml") -> dict:
    """Load configuration from YAML file."""
    if not os.path.exists(config_path):
        logger.error(f"Config file not found: {config_path}")
        sys.exit(1)

    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)

    return config


def merge_applied_companies(config: dict, base_path: str) -> dict:
    """Merge companies from applied positions file into config."""
    applied_file = config.get("applied_positions_file", "")
    if applied_file:
        full_path = os.path.join(base_path, applied_file) if not os.path.isabs(applied_file) else applied_file
        applied_companies = load_applied_companies(full_path)

        if applied_companies:
            logger.info(f"Found {len(applied_companies)} companies from applied positions file")
            # Add as additional companies (won't have specific scrapers but logged for reference)
            existing_additional = config.get("additional_companies", [])
            config["additional_companies"] = list(set(existing_additional + applied_companies))

    return config


def load_company_api_file(config: dict, base_path: str) -> dict:
    """Load company API information from 公司API.xlsx file.

    Returns:
        dict: Updated config with companies from API file
    """
    api_file = os.path.join(base_path, "公司API.xlsx")
    if not os.path.exists(api_file):
        logger.info("公司API.xlsx not found, skipping API file load")
        return config

    try:
        df = pd.read_excel(api_file)
        logger.info(f"Loaded {len(df)} companies from 公司API.xlsx")

        existing_companies = config.get("companies", {})
        added = 0
        skipped_types = []

        for _, row in df.iterrows():
            company_name = str(row.get("公司名称 (Company Name)", "")).strip()
            ats_type = str(row.get("招聘系统 (ATS Type)", "")).strip()
            api_url = str(row.get("建议的 API 路径 / 职位页面 URL", "")).strip()

            if not company_name or not api_url:
                continue

            # Skip entries without usable APIs
            if "需视" in api_url or "OData" in api_url or "需Token" in api_url or api_url == "nan":
                skipped_types.append(company_name)
                continue

            # Skip custom/internal systems we can't scrape
            if ats_type in ["Custom", "Custom/Internal", "Proprietary", "各子公司不同", "各子公司独立",
                           "Taleo/Custom", "Custom/BrassRing", "iCIMS", "Oracle Cloud", "ADP",
                           "Oracle/Custom", "Jobvite", "Avature"]:
                skipped_types.append(company_name)
                continue

            # Create a normalized key
            company_key = company_name.replace(" ", "").replace(".", "").replace(",", "")

            # Skip if already in config
            if company_key in existing_companies:
                continue

            # Check by name too
            existing_names = {v.get("name", "").lower() for v in existing_companies.values()}
            if company_name.lower() in existing_names:
                continue

            # Determine scraper type
            scraper_type = ""
            if "Workday" in ats_type or "myworkdayjobs.com" in api_url.lower():
                scraper_type = "workday"
            elif "Greenhouse" in ats_type or "greenhouse.io" in api_url.lower():
                scraper_type = "greenhouse"
            elif "Lever" in ats_type or "lever.co" in api_url.lower():
                scraper_type = "lever"
            elif "SmartRecruiters" in ats_type or "smartrecruiters.com" in api_url.lower():
                scraper_type = "smartrecruiters"
            elif "eightfold" in ats_type.lower() or "eightfold.ai" in api_url.lower():
                scraper_type = "eightfold"
            elif "SuccessFactors" in ats_type:
                # SuccessFactors typically needs OData, skip for now
                skipped_types.append(company_name)
                continue
            else:
                # Unknown type, skip
                skipped_types.append(company_name)
                continue

            existing_companies[company_key] = {
                "name": company_name,
                "type": scraper_type,
                "api_url": api_url,
            }
            added += 1

        config["companies"] = existing_companies
        logger.info(f"Added {added} companies from 公司API.xlsx")
        logger.info(f"Skipped {len(skipped_types)} companies with unsupported ATS types")
        logger.info(f"Total companies to search: {len(existing_companies)}")

    except Exception as e:
        logger.error(f"Error loading 公司API.xlsx: {e}")

    return config


def load_company_list(config: dict, base_path: str) -> tuple:
    """Load and merge companies from external company list file.

    Returns:
        tuple: (updated_config, unmatched_companies)
    """
    company_list_file = config.get("company_list_file", "")
    if not company_list_file:
        return config, []

    full_path = os.path.join(base_path, company_list_file) if not os.path.isabs(company_list_file) else company_list_file

    if not os.path.exists(full_path):
        logger.warning(f"Company list file not found: {full_path}")
        return config, []

    # Load companies from Excel
    companies = load_companies_from_excel(full_path)
    if not companies:
        return config, []

    logger.info(f"Loaded {len(companies)} companies from {company_list_file}")

    # Auto-discover career APIs
    new_configs, unmatched = generate_company_configs(companies)

    # Merge with existing companies config
    existing_companies = config.get("companies", {})
    existing_keys = {k.lower() for k in existing_companies.keys()}
    existing_names = {v.get("name", "").lower() for v in existing_companies.values()}

    added = 0
    for key, company_config in new_configs.items():
        if key.lower() not in existing_keys and company_config.get("name", "").lower() not in existing_names:
            existing_companies[key] = company_config
            added += 1

    config["companies"] = existing_companies
    logger.info(f"Added {added} new companies with discoverable career APIs")
    logger.info(f"Total companies to search: {len(existing_companies)}")
    logger.info(f"Companies without discoverable APIs: {len(unmatched)}")

    return config, unmatched


def save_unsearchable_companies(companies: list, output_path: str):
    """Save unsearchable companies to an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Unsearchable Companies"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Headers
    ws["A1"] = "Company Name"
    ws["B1"] = "Status"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["B1"].font = header_font
    ws["B1"].fill = header_fill

    # Data
    for i, company in enumerate(sorted(companies), start=2):
        ws[f"A{i}"] = company
        ws[f"B{i}"] = "No career API found"

    # Adjust column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 25

    wb.save(output_path)
    logger.info(f"Saved {len(companies)} unsearchable companies to {output_path}")


def save_categorized_companies(categories: dict, output_path: str):
    """Save categorized company search results to an Excel file with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    fills = {
        "no_api_found": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),  # Red
        "no_matching_jobs": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),  # Orange
        "api_error": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),  # Yellow
    }
    status_labels = {
        "no_api_found": "No career API discoverable - requires manual search",
        "no_matching_jobs": "API works but no matching intern/co-op positions found",
        "api_error": "API returned error or connection failed",
    }
    sheet_names = {
        "no_api_found": "No API Found",
        "no_matching_jobs": "No Matching Jobs",
        "api_error": "API Errors",
    }

    first_sheet = True
    for category, companies in categories.items():
        if not companies:
            continue

        if first_sheet:
            ws = wb.active
            ws.title = sheet_names.get(category, category)
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_names.get(category, category))

        # Headers
        ws["A1"] = "Company Name"
        ws["B1"] = "Status"
        ws["A1"].font = header_font
        ws["A1"].fill = fills.get(category, PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"))
        ws["B1"].font = header_font
        ws["B1"].fill = fills.get(category, PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"))

        # Data
        status = status_labels.get(category, category)
        for i, company in enumerate(sorted(companies), start=2):
            ws[f"A{i}"] = company
            ws[f"B{i}"] = status

        # Adjust column widths
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 60

    # If no sheets were created, add a summary
    if first_sheet:
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = "All companies returned results successfully!"

    wb.save(output_path)
    total = sum(len(c) for c in categories.values())
    logger.info(f"Saved {total} categorized companies to {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Search for matching intern/co-op positions at major tech companies"
    )
    parser.add_argument(
        "--days", "-d",
        type=int,
        default=None,
        help="Number of days to look back (overrides config)"
    )
    parser.add_argument(
        "--config", "-c",
        type=str,
        default="config.yaml",
        help="Path to config file"
    )
    parser.add_argument(
        "--no-similarity",
        action="store_true",
        help="Skip similarity matching (faster, no model loading)"
    )
    parser.add_argument(
        "--output-dir", "-o",
        type=str,
        default=None,
        help="Output directory (overrides config)"
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable verbose output"
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Determine base path
    base_path = os.path.dirname(os.path.abspath(__file__))
    os.chdir(base_path)

    print("\n" + "=" * 60)
    print("Job Seeking Tool")
    print("=" * 60)
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # Load configuration
    config_path = args.config if os.path.isabs(args.config) else os.path.join(base_path, args.config)
    logger.info(f"Loading config from: {config_path}")
    config = load_config(config_path)

    # Override days if specified
    if args.days:
        config["days_lookback"] = args.days

    days_back = config.get("days_lookback", 7)
    print(f"Looking for jobs posted in the last {days_back} days")
    print()

    # Merge applied companies
    config = merge_applied_companies(config, base_path)

    # Load additional companies from company list file
    config, unsearchable_companies = load_company_list(config, base_path)

    # Load companies from API file (公司API.xlsx) - this has verified API URLs
    config = load_company_api_file(config, base_path)

    # Fetch jobs from all companies
    print("Fetching jobs from company career pages...")
    print("-" * 40)

    jobs, search_results = fetch_all_jobs(config)

    print("-" * 40)
    print(f"Total matching jobs found: {len(jobs)}")
    print()

    if not jobs:
        print("No matching jobs found. Try adjusting your search criteria.")

    # Compute similarity scores
    similarity_results = []

    if not args.no_similarity:
        print("Computing resume similarity scores...")
        print("-" * 40)

        try:
            matcher = create_matcher(config, base_path)

            if matcher.resumes:
                for i, job in enumerate(jobs):
                    result = matcher.compute_similarity(job.title, job.description)
                    similarity_results.append(result)

                    if (i + 1) % 10 == 0:
                        print(f"  Processed {i + 1}/{len(jobs)} jobs...")

                print(f"  Completed similarity matching for {len(jobs)} jobs")
            else:
                print("  Warning: No resumes loaded, skipping similarity matching")
                similarity_results = [None] * len(jobs)

        except Exception as e:
            logger.error(f"Error during similarity matching: {e}")
            print(f"  Warning: Similarity matching failed: {e}")
            print("  Continuing without similarity scores...")
            similarity_results = [None] * len(jobs)
    else:
        print("Skipping similarity matching (--no-similarity flag)")
        similarity_results = [None] * len(jobs)

    print()

    # Generate output
    output_dir = args.output_dir or config.get("output", {}).get("directory", "output")
    output_dir = output_dir if os.path.isabs(output_dir) else os.path.join(base_path, output_dir)

    print("Generating Excel output...")
    print("-" * 40)

    output_path = generate_excel_output(jobs, similarity_results, config, output_dir)

    print()
    print("=" * 60)
    print("COMPLETE!")
    print("=" * 60)
    print(f"Output file: {output_path}")
    print(f"Total jobs: {len(jobs)}")

    # Print summary by company
    company_counts = {}
    for job in jobs:
        company_counts[job.company] = company_counts.get(job.company, 0) + 1

    print("\nJobs by company:")
    for company, count in sorted(company_counts.items(), key=lambda x: -x[1]):
        print(f"  {company}: {count}")

    # Print summary by recommended resume
    if similarity_results and any(r for r in similarity_results):
        resume_counts = {}
        for result in similarity_results:
            if result and result.recommended_resume:
                resume_counts[result.recommended_resume] = resume_counts.get(result.recommended_resume, 0) + 1

        if resume_counts:
            print("\nJobs by recommended resume:")
            for resume, count in sorted(resume_counts.items(), key=lambda x: -x[1]):
                print(f"  {resume}: {count}")

    # Output company search results by category
    print(f"\n{'=' * 60}")
    print("COMPANY SEARCH RESULTS SUMMARY")
    print("=" * 60)

    # 1. Companies with matching jobs found
    if search_results.get("success"):
        print(f"\n[SUCCESS] Companies with matching jobs: {len(search_results['success'])}")

    # 2. Companies with API but no matching jobs
    no_jobs = search_results.get("no_matching_jobs", [])
    if no_jobs:
        print(f"\n[NO MATCHING JOBS] API worked but no matching intern positions: {len(no_jobs)}")
        for company in sorted(no_jobs)[:20]:  # Show first 20
            print(f"  - {company}")
        if len(no_jobs) > 20:
            print(f"  ... and {len(no_jobs) - 20} more")

    # 3. Companies with API errors
    api_errors = search_results.get("api_error", [])
    if api_errors:
        print(f"\n[API ERROR] API returned error or failed: {len(api_errors)}")
        for company in sorted(api_errors)[:20]:  # Show first 20
            print(f"  - {company}")
        if len(api_errors) > 20:
            print(f"  ... and {len(api_errors) - 20} more")

    # 4. Companies with no discoverable API
    if unsearchable_companies:
        print(f"\n[NO API FOUND] No discoverable career API: {len(unsearchable_companies)}")
        for company in sorted(unsearchable_companies)[:30]:  # Show first 30
            print(f"  - {company}")
        if len(unsearchable_companies) > 30:
            print(f"  ... and {len(unsearchable_companies) - 30} more")

    # Save all categories to Excel file
    all_categorized = {
        "no_api_found": unsearchable_companies,
        "no_matching_jobs": no_jobs,
        "api_error": api_errors,
    }
    categorized_file = os.path.join(output_dir, f"company_search_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    save_categorized_companies(all_categorized, categorized_file)
    print(f"\nCompany search results saved to: {categorized_file}")

    print()


if __name__ == "__main__":
    main()
